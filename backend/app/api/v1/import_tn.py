from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional
import io

from app.core.database import get_db
from app.core.config import settings as app_settings
from app.models import TnEntry, User
from app.schemas.common import ApiResponse
from app.api.v1.auth import get_current_user

router = APIRouter(prefix="/import", tags=["Import"])

EXCEL_COLUMNS = [
    ("name", "Name", "Unique identifier, e.g. ISLEEU-1", True),
    ("family", "Family", "e.g. PIF-Harbinger, Tc1-Mariner", True),
    ("tn_group", "Group", "e.g. ISL2EU", True),
    ("synonyms", "Synonyms", "Alternative names, separated by semicolons", False),
    ("isoform", "Isoform", "Isoform identifier", False),
    ("accession_number", "Accession Number", "Genome accession, e.g. GCA_943735915_1", False),
    ("origin", "Origin", "Source species name", False),
    ("mge_type", "MGE Type", "TE / MITE / LARD / TRIM", False),
    ("related_elements", "Related Elements", "Related transposon elements", False),
    ("length", "Length", "Total length in bp (integer)", False),
    ("ir", "IR", "Inverted repeat match, e.g. 12/13", False),
    ("dr", "DR", "Direct repeat length in bp (integer)", False),
    ("orf", "ORF", "ORF sizes, e.g. 447/308", False),
    ("irl", "IRL", "Left inverted repeat sequence", False),
    ("irr", "IRR", "Right inverted repeat sequence", False),
    ("left_flank", "Left Flank", "Left flanking sequence", False),
    ("right_flank", "Right Flank", "Right flanking sequence", False),
    ("transposition", "Transposition", "e.g. Cut-and-paste, Rolling-circle", False),
    ("direct_repeat", "Direct Repeat", "TSD sequence", False),
    ("dna_sequence", "DNA Sequence", "Full nucleotide sequence (ATCG)", False),
    ("orf1_name", "ORF1 Name", "ORF1 identifier", False),
    ("orf1_length", "ORF1 Length", "ORF1 length in bp (integer)", False),
    ("orf1_begin", "ORF1 Begin", "ORF1 start position (integer)", False),
    ("orf1_end", "ORF1 End", "ORF1 end position (integer)", False),
    ("orf1_strand", "ORF1 Strand", "+ or -", False),
    ("orf1_fusion_orf", "ORF1 Fusion ORF", "Yes or No", False),
    ("orf1_function", "ORF1 Function", "e.g. Transposase", False),
    ("orf1_chemistry", "ORF1 Chemistry", "e.g. DDE, DDD, Y1", False),
    ("orf1_sequence", "ORF1 Sequence", "ORF1 protein sequence", False),
    ("orf2_name", "ORF2 Name", "ORF2 identifier", False),
    ("orf2_length", "ORF2 Length", "ORF2 length in bp (integer)", False),
    ("orf2_begin", "ORF2 Begin", "ORF2 start position (integer)", False),
    ("orf2_end", "ORF2 End", "ORF2 end position (integer)", False),
    ("orf2_strand", "ORF2 Strand", "+ or -", False),
    ("orf2_fusion_orf", "ORF2 Fusion ORF", "Yes or No", False),
    ("orf2_function", "ORF2 Function", "e.g. Yqaj", False),
    ("orf2_chemistry", "ORF2 Chemistry", "e.g. DDE", False),
    ("orf2_sequence", "ORF2 Sequence", "ORF2 protein sequence", False),
]


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Permission denied")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "euTnDB Import Template"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    required_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    desc_font = Font(name="Calibri", italic=True, color="5F6368", size=10)
    thin_border = Border(
        left=Side(style="thin", color="DADCE0"),
        right=Side(style="thin", color="DADCE0"),
        top=Side(style="thin", color="DADCE0"),
        bottom=Side(style="thin", color="DADCE0"),
    )

    ws.append([col[1] for col in EXCEL_COLUMNS])
    for col_idx, (field, header, desc, required) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.append([col[2] for col in EXCEL_COLUMNS])
    for col_idx, (field, header, desc, required) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = desc_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.append([col[1] for col in EXCEL_COLUMNS])
    for col_idx, (field, header, desc, required) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.border = thin_border
        if required:
            cell.fill = required_fill

    example_data = {
        "name": "ISLEEU-1",
        "family": "PIF-Harbinger",
        "tn_group": "ISL2EU",
        "synonyms": "",
        "isoform": "",
        "accession_number": "GCA_943735915_1",
        "origin": "Pherbina coryleti",
        "mge_type": "TE",
        "related_elements": "",
        "length": 5432,
        "ir": "12/13",
        "dr": 0,
        "orf": "447/308",
        "irl": "",
        "irr": "",
        "left_flank": "",
        "right_flank": "",
        "transposition": "Cut-and-paste",
        "direct_repeat": "-",
        "dna_sequence": "AATAGGGT...",
        "orf1_name": "",
        "orf1_length": 1344,
        "orf1_begin": 995,
        "orf1_end": 2338,
        "orf1_strand": "+",
        "orf1_fusion_orf": "No",
        "orf1_function": "Transposase",
        "orf1_chemistry": "DDE",
        "orf1_sequence": "",
        "orf2_name": "",
        "orf2_length": 927,
        "orf2_begin": 4277,
        "orf2_end": 3351,
        "orf2_strand": "-",
        "orf2_fusion_orf": "No",
        "orf2_function": "Yqaj",
        "orf2_chemistry": "",
        "orf2_sequence": "",
    }
    ws.append([example_data.get(col[0], "") for col in EXCEL_COLUMNS])
    for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.border = thin_border

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    for col_idx, (field, header, desc, required) in enumerate(EXCEL_COLUMNS, 1):
        max_len = max(len(header), len(desc), 15)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(EXCEL_COLUMNS))}1"
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=euTnDB_Import_Template.xlsx"}
    )


@router.post("/excel", response_model=ApiResponse)
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Permission denied")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload .xlsx or .xls file")

    import openpyxl

    content = await file.read()
    if len(content) > app_settings.MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    ws = wb.active

    header_row = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        header_row.append(str(cell).strip() if cell else "")

    header_map = {}
    for col_idx, header_text in enumerate(header_row):
        for field, header, desc, required in EXCEL_COLUMNS:
            if header_text.lower() == header.lower():
                header_map[field] = col_idx
                break

    if "name" not in header_map or "family" not in header_map or "tn_group" not in header_map:
        missing = []
        if "name" not in header_map:
            missing.append("Name")
        if "family" not in header_map:
            missing.append("Family")
        if "tn_group" not in header_map:
            missing.append("Group")
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}. Please use the provided template.")

    created_count = 0
    skipped_count = 0
    error_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        row_data = {}
        for field, col_idx in header_map.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None:
                    row_data[field] = val

        name_val = row_data.get("name")
        if not name_val or str(name_val).strip() == "":
            continue

        name_val = str(name_val).strip()

        if str(row_data.get("name", "")).strip() == "" and all(
            str(row_data.get(f, "")).strip() == ""
            for f in ["family", "tn_group", "origin", "dna_sequence"]
        ):
            continue

        existing = await db.execute(
            select(TnEntry).where(TnEntry.name == name_val)
        )
        if existing.scalar_one_or_none():
            skipped_count += 1
            error_rows.append({"row": row_idx, "name": name_val, "error": "Name already exists"})
            continue

        family_val = str(row_data.get("family", "")).strip()
        group_val = str(row_data.get("tn_group", "")).strip()

        if not family_val or not group_val:
            skipped_count += 1
            error_rows.append({"row": row_idx, "name": name_val, "error": "Missing required field: family or group"})
            continue

        entry_data = {"name": name_val, "family": family_val, "tn_group": group_val}

        int_fields = ["length", "dr", "orf1_length", "orf1_begin", "orf1_end",
                       "orf2_length", "orf2_begin", "orf2_end"]
        str_fields = ["synonyms", "isoform", "accession_number", "origin", "mge_type",
                       "related_elements", "ir", "orf", "irl", "irr", "left_flank",
                       "right_flank", "transposition", "direct_repeat", "dna_sequence",
                       "orf1_name", "orf1_strand", "orf1_fusion_orf", "orf1_function",
                       "orf1_chemistry", "orf1_sequence", "orf2_name", "orf2_strand",
                       "orf2_fusion_orf", "orf2_function", "orf2_chemistry", "orf2_sequence"]

        for f in int_fields:
            if f in row_data and row_data[f] is not None:
                try:
                    entry_data[f] = int(row_data[f])
                except (ValueError, TypeError):
                    pass

        for f in str_fields:
            if f in row_data and row_data[f] is not None:
                val = str(row_data[f]).strip()
                if val:
                    entry_data[f] = val

        entry = TnEntry(**entry_data, submitted_by=current_user.id)
        db.add(entry)
        created_count += 1

    await db.commit()
    wb.close()

    return ApiResponse(data={
        "created": created_count,
        "skipped": skipped_count,
        "errors": error_rows[:20],
        "total_rows": created_count + skipped_count
    })
